"""从 assets/静态数据/话题热度分析_v9.xlsx 中的"异动时间线"和"内容异动KPI"sheet
生成 src/data/contentAnomalyStatic.js"""
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "assets" / "静态数据" / "话题热度分析_v9.xlsx"
OUT = ROOT / "src" / "data" / "contentAnomalyStatic.js"


def to_number(value):
    if value in (None, "", "-"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_int(value):
    if value in (None, "", "-"):
        return 0
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return 0


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    # ─── 读取 内容异动KPI ──────────────────────────────────────────────
    kpi_ws = wb.worksheets[5]  # sheet index 5
    kpi = {}
    for row in kpi_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        label = clean_text(row[0])
        value = to_int(row[1])
        if "暴涨话题" in label:
            kpi["surgeTopic"] = value
        elif "新出话题" in label:
            kpi["newTopic"] = value
        elif "高贴合" in label:
            kpi["highAlignInstitution"] = value
        elif "爆款帖子" in label:
            kpi["viralPosts"] = value

    # ─── 读取 异动时间线 ────────────────────────────────────────────────
    # 列：类型 | 话题名称 | 近期占比% | 前期占比% | 热度变化% | 近期内容数 | 前期内容数 | 首现天数 | 主发平台
    tl_ws = wb["异动时间线"]
    timeline = []
    for row in tl_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        type_raw = clean_text(row[0])
        topic = clean_text(row[1])
        current_share = to_number(row[2]) or 0.0
        previous_share = to_number(row[3]) or 0.0
        change_raw = to_number(row[4])        # '-' → None for new topics
        current_count = to_int(row[5])
        previous_count = to_int(row[6])
        first_seen_days = to_int(row[7])
        main_platform = clean_text(row[8])

        item_type = "new" if "新出" in type_raw else "surge"

        timeline.append({
            "type": item_type,
            "topic": topic,
            "currentShare": round(current_share, 2),
            "previousShare": round(previous_share, 2),
            "change": round(change_raw, 1) if change_raw is not None else None,
            "currentCount": current_count,
            "previousCount": previous_count,
            "firstSeenDays": first_seen_days,
            "mainPlatform": main_platform,
        })

    wb.close()

    # ─── 写出 JS ───────────────────────────────────────────────────────
    lines = [
        "// 来源：assets/静态数据/话题热度分析_v9.xlsx（异动时间线 + 内容异动KPI）",
        "// 重新生成：python scripts/emit_anomaly_static.py",
        "",
        "export const ANOMALY_KPI = {",
        f"  surgeTopic: {kpi.get('surgeTopic', 0)},",
        f"  newTopic: {kpi.get('newTopic', 0)},",
        f"  highAlignInstitution: {kpi.get('highAlignInstitution', 0)},",
        f"  viralPosts: {kpi.get('viralPosts', 0)},",
        "};",
        "",
        "export const ANOMALY_TIMELINE = [",
    ]
    for i, item in enumerate(timeline):
        sep = "," if i < len(timeline) - 1 else ""
        change_str = str(item["change"]) if item["change"] is not None else "null"
        lines.append(
            f"  {{"
            f' type: {json.dumps(item["type"], ensure_ascii=False)},'
            f' topic: {json.dumps(item["topic"], ensure_ascii=False)},'
            f' currentShare: {item["currentShare"]},'
            f' previousShare: {item["previousShare"]},'
            f' change: {change_str},'
            f' currentCount: {item["currentCount"]},'
            f' previousCount: {item["previousCount"]},'
            f' firstSeenDays: {item["firstSeenDays"]},'
            f' mainPlatform: {json.dumps(item["mainPlatform"], ensure_ascii=False)}'
            f" }}{sep}"
        )
    lines.append("];")
    lines.append("")

    OUT.write_text("\n".join(lines), encoding="utf-8")
    surge_count = sum(1 for t in timeline if t["type"] == "surge")
    new_count = sum(1 for t in timeline if t["type"] == "new")
    print(f"KPI: {kpi}")
    print(f"Timeline: {len(timeline)} rows ({surge_count} surge, {new_count} new) -> {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
