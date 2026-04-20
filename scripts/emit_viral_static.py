"""从 assets/静态数据/爆款内容墙.xlsx 生成 src/data/viralPostsWallStatic.js"""
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "assets" / "静态数据" / "爆款内容墙.xlsx"
OUT = ROOT / "src" / "data" / "viralPostsWallStatic.js"


def map_channel(ch):
    m = {
        "小红书": "小红书",
        "蚂蚁财富": "蚂蚁财富号",
        "微信公众号": "微信公众号",
        "雪球": "雪球",
        "抖音": "抖音",
    }
    return m.get(str(ch).strip(), "微信公众号")


def to_int(v):
    if v is None or v == "-":
        return 0
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


def fmt_interaction(row):
    s = sum(to_int(row[i]) for i in range(6, 10))
    if s >= 10000:
        return f"{s / 10000:.1f}万"
    return str(s)


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    out = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        ch, account, title = row[0], row[1], row[2]
        url = row[3] if len(row) > 3 else ""
        ctype = row[4] if len(row) > 4 else ""
        pub = row[5] if len(row) > 5 else ""
        title_str = (
            (title or "").strip()
            if title and str(title).strip() != "-"
            else "（无标题）"
        )
        tags = []
        if ctype and str(ctype).strip() and str(ctype) != "-":
            tags.append(f"#{ctype}")
        if ch:
            tags.append(f"#{ch}")
        if not tags:
            tags = ["#爆款"]
        out.append(
            {
                "title": title_str,
                "institution": str(account or "").strip() or "未知账号",
                "platform": map_channel(ch),
                "interactions": fmt_interaction(row),
                "tags": tags[:6],
                "url": url or "",
                "contentType": str(ctype or ""),
                "publishedAt": str(pub or ""),
            }
        )

    lines = [
        "// 来源：assets/静态数据/爆款内容墙.xlsx",
        "// 表头：渠道、账号名称、标题、资源链接、内容类型、发布时间(day)、点赞数、分享数、评论数、收藏数",
        "// 重新生成：python scripts/emit_viral_static.py",
        "export const VIRAL_POSTS = [",
    ]
    for i, o in enumerate(out):
        block = "  {\n"
        block += f"    title: {json.dumps(o['title'], ensure_ascii=False)},\n"
        block += f"    institution: {json.dumps(o['institution'], ensure_ascii=False)},\n"
        block += f"    platform: {json.dumps(o['platform'], ensure_ascii=False)},\n"
        block += f"    interactions: {json.dumps(o['interactions'], ensure_ascii=False)},\n"
        block += f"    tags: {json.dumps(o['tags'], ensure_ascii=False)},\n"
        block += f"    url: {json.dumps(o.get('url', ''), ensure_ascii=False)},\n"
        block += f"    contentType: {json.dumps(o.get('contentType', ''), ensure_ascii=False)},\n"
        block += f"    publishedAt: {json.dumps(o.get('publishedAt', ''), ensure_ascii=False)},\n"
        block += "  }" + ("," if i < len(out) - 1 else "")
        lines.append(block)
    lines.append("];")
    lines.append("")

    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"written {len(out)} rows -> {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
