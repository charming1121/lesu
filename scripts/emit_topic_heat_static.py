"""从 assets/静态数据/话题热度分析_v10.xlsx 生成 src/data/topicHeatAnalysisStatic.js"""
import json
import math
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "assets" / "静态数据" / "话题热度分析_v10.xlsx"
OUT = ROOT / "src" / "data" / "topicHeatAnalysisStatic.js"

PLATFORMS = ["小红书", "蚂蚁财富号", "微信公众号", "雪球"]
PLATFORM_HEADER_ALIASES = {
    "小红书": "小红书",
    "蚂蚁财富号": "蚂蚁财富",
    "微信公众号": "微信公众号",
    "雪球": "雪球",
}


def normalize_header(value):
    return str(value or "").strip().replace(" ", "")


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def to_int(value):
    if value in (None, "", "-"):
        return 0
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return 0


def to_number(value):
    if value in (None, "", "-"):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def to_percent(value):
    numeric = round(to_number(value), 1)
    if math.isclose(numeric, round(numeric)):
        return int(round(numeric))
    return numeric


def percent_mix(entries):
    non_zero_entries = [(name, count) for name, count in entries if count > 0]
    if not non_zero_entries:
        return []

    total = sum(count for _, count in non_zero_entries)
    raw_values = []
    floor_sum = 0
    for name, count in non_zero_entries:
        raw = count * 100 / total
        floor_value = math.floor(raw)
        raw_values.append(
            {
                "name": name,
                "raw": raw,
                "value": floor_value,
                "fraction": raw - floor_value,
            }
        )
        floor_sum += floor_value

    remainder = 100 - floor_sum
    for item in sorted(raw_values, key=lambda current: current["fraction"], reverse=True)[:remainder]:
        item["value"] += 1

    return [
        {"name": item["name"], "value": item["value"]}
        for item in sorted(raw_values, key=lambda current: (-current["value"], current["name"]))
    ]


def ensure_detail(detail_map, topic):
    if topic not in detail_map:
        detail_map[topic] = {
            "topic": topic,
            "totalCount": 0,
            "totalShare": 0,
            "platformBreakdown": {
                platform: {"count": 0, "share": 0} for platform in PLATFORMS
            },
            "institutions": [],
            "products": [],
        }
    return detail_map[topic]


def build_header_index(ws):
    headers = [normalize_header(cell) for cell in next(ws.iter_rows(values_only=True))]
    return {header: index for index, header in enumerate(headers)}


def read_overall_sheet(ws, detail_map):
    header_index = build_header_index(ws)
    rows = list(ws.iter_rows(values_only=True))[1:]

    overall_list = []
    for row in rows:
        topic = clean_text(row[header_index["标准话题"]])
        if not topic:
            continue

        detail = ensure_detail(detail_map, topic)
        detail["totalCount"] = to_int(row[header_index["全平台内容数"]])
        detail["totalShare"] = to_percent(row[header_index["热度占比%"]])

        for platform in PLATFORMS:
            header_name = PLATFORM_HEADER_ALIASES[platform]
            detail["platformBreakdown"][platform]["count"] = to_int(row[header_index[header_name]])

        overall_list.append(
            {
                "rank": to_int(row[header_index["排名"]]),
                "topic": topic,
                "count": detail["totalCount"],
                "share": detail["totalShare"],
            }
        )

    return sorted(overall_list, key=lambda item: (item["rank"], -item["count"], item["topic"]))


def read_penetration_sheet(ws, detail_map):
    header_index = build_header_index(ws)
    rows = list(ws.iter_rows(values_only=True))[1:]
    platform_lists = defaultdict(list)

    for row in rows:
        topic = clean_text(row[header_index["标准话题"]])
        if not topic:
            continue

        detail = ensure_detail(detail_map, topic)
        detail["totalCount"] = max(detail["totalCount"], to_int(row[header_index["全平台内容数"]]))

        for platform in PLATFORMS:
            source_prefix = PLATFORM_HEADER_ALIASES[platform]
            share_key = normalize_header(f"{source_prefix}_渗透率%")
            count_key = normalize_header(f"{source_prefix}_内容数")
            count = to_int(row[header_index[count_key]])
            share = to_percent(row[header_index[share_key]])
            detail["platformBreakdown"][platform] = {"count": count, "share": share}
            platform_lists[platform].append({"topic": topic, "count": count, "share": share})

    ranking_by_platform = {}
    for platform in PLATFORMS:
        sorted_rows = sorted(
            [item for item in platform_lists[platform] if item["count"] > 0],
            key=lambda item: (-item["count"], -to_number(item["share"]), item["topic"]),
        )
        ranking_by_platform[platform] = [
            {
                "rank": index + 1,
                "topic": item["topic"],
                "count": item["count"],
                "share": item["share"],
            }
            for index, item in enumerate(sorted_rows)
        ]

    return ranking_by_platform


def read_institution_sheet(ws, detail_map):
    header_index = build_header_index(ws)
    rows = list(ws.iter_rows(values_only=True))[1:]

    for row in rows:
        topic = clean_text(row[header_index["标准话题"]])
        institution = clean_text(row[header_index["机构名称"]])
        if not topic or not institution:
            continue

        detail = ensure_detail(detail_map, topic)
        participation_type = clean_text(row[header_index["参与类型"]])
        detail["institutions"].append(
            {
                "institution": institution,
                "value": to_percent(row[header_index["话题内占比%"]]),
                "count": to_int(row[header_index["发布内容数"]]),
                "primary": "核心" in participation_type,
                "participationType": participation_type,
            }
        )

    for detail in detail_map.values():
        detail["institutions"] = sorted(
            detail["institutions"],
            key=lambda item: (-to_number(item["value"]), -item["count"], item["institution"]),
        )


def read_product_sheet(ws, detail_map):
    header_index = build_header_index(ws)
    rows = list(ws.iter_rows(values_only=True))[1:]

    for row in rows:
        topic = clean_text(row[header_index["标准话题"]])
        if not topic:
            continue

        institution = clean_text(row[header_index["机构名称"]]) or clean_text(row[header_index["管理公司"]]) or "未命名机构"
        product_name = clean_text(row[header_index["产品名称"]]) or institution
        product_code = clean_text(row[header_index["产品代码"]])
        count = to_int(row[header_index["关联内容数"]])
        ratio = to_percent(row[header_index["话题内推品占比%"]])
        platform_counts = [
            ("小红书", to_int(row[header_index[normalize_header("小红书_内容数")]])),
            ("蚂蚁财富号", to_int(row[header_index[normalize_header("蚂蚁财富_内容数")]])),
            ("微信公众号", to_int(row[header_index[normalize_header("微信公众号_内容数")]])),
            ("雪球", to_int(row[header_index[normalize_header("雪球_内容数")]])),
        ]

        detail = ensure_detail(detail_map, topic)
        detail["products"].append(
            {
                "code": product_code,
                "product": product_name,
                "institution": institution,
                "ratio": ratio,
                "count": count,
                "platformMix": percent_mix(platform_counts),
            }
        )

    for detail in detail_map.values():
        detail["products"] = sorted(
            detail["products"],
            key=lambda item: (-to_number(item["ratio"]), -item["count"], item["product"]),
        )


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    detail_map = {}

    overall_list = read_overall_sheet(wb["话题热度榜"], detail_map)
    platform_lists = read_penetration_sheet(wb["渠道渗透率"], detail_map)
    read_institution_sheet(wb["机构×话题占比"], detail_map)
    read_product_sheet(wb["机构推品列表"], detail_map)
    wb.close()

    market_topic_heat_by_platform = {"全平台": overall_list}
    market_topic_heat_by_platform.update(platform_lists)

    lines = [
        "// 来源：assets/静态数据/话题热度分析_v10.xlsx",
        "// 工作表：话题热度榜、渠道渗透率、机构×话题占比、机构推品列表",
        "// 重新生成：python scripts/emit_topic_heat_static.py",
        "export const MARKET_TOPIC_HEAT_BY_PLATFORM = " + json.dumps(market_topic_heat_by_platform, ensure_ascii=False, indent=2) + ";",
        "",
        "export const MARKET_TOPIC_DETAILS = " + json.dumps(detail_map, ensure_ascii=False, indent=2) + ";",
        "",
    ]
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"written {len(detail_map)} topics -> {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()